from pywhatkit import sendwhatmsg_instantly
from openpyxl import load_workbook

wb = load_workbook('nums.xlsx')
sh1 = wb['Sheet1'
]
for i in range(1,6):
    
    Numbr = '+20'+str(sh1.cell(i,1).value)
    sendwhatmsg_instantly(Numbr,"AUTOMATED PLEASE IGNORE",20,True,10)
