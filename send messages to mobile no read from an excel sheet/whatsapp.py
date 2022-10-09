import pywhatkit
from openpyxl import load_workbook
wb = load_workbook('numbers.xlsx')
sh1 = wb['Sheet1']
k = 0
for i in range(1,4):
    Numbr = '+20'+str(sh1.cell(i,1).value)
    pywhatkit.sendwhatmsg(Numbr,"PLEASE IGNORE",22,46+k,10,True,10)
    k+= 3